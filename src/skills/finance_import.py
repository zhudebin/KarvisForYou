# -*- coding: utf-8 -*-
"""
Skill: finance.import
iCost xlsx 文件扫描、解析、去重、合并到 finance_data.json。

触发方式：
  - 手动：用户说「导入 iCost 数据」
  - 自动：月报生成前自动调用

V12 移植：handler 签名 (params, state, ctx)，IO 通过 ctx.IO。
"""
import sys
import io
from finance_utils import (
    load_finance_data, save_finance_data, parse_date, parse_amount, _log
)


# iCost xlsx 中需要提取的列（按名称匹配，顺序无关）
_TARGET_COLUMNS = ["日期", "类型", "金额", "一级分类", "二级分类", "备注"]


def _bill_hash(bill):
    """生成账单去重指纹：日期+金额+一级分类+二级分类+备注。"""
    parts = [
        str(bill.get("日期", "")),
        str(bill.get("金额", "")),
        str(bill.get("一级分类", "")),
        str(bill.get("二级分类", "")),
        str(bill.get("备注", "")),
    ]
    return "_".join(parts)


def _parse_xlsx_bytes(data_bytes, filename=""):
    """解析 iCost xlsx 二进制数据，返回记录列表。

    Args:
        data_bytes: bytes — xlsx 文件内容
        filename: str — 文件名（用于日志）

    Returns:
        (list[dict], str|None) — (解析出的记录列表, 错误信息)
    """
    try:
        import openpyxl
    except ImportError:
        return [], "服务端缺少 openpyxl 库，无法解析 xlsx 文件"

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data_bytes), read_only=True)
    except Exception as e:
        return [], f"无法打开 xlsx 文件 {filename}: {e}"

    # Sheet 选择：优先找「收支账单」，否则取第一个
    sheet_name = None
    if "收支账单" in wb.sheetnames:
        sheet_name = "收支账单"
    else:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # 读取所有行
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return [], f"Sheet「{sheet_name}」数据为空（只有表头或没有数据）"

    # 解析表头，建立列名 → 索引映射
    headers = [str(h).strip() if h else "" for h in rows[0]]
    col_map = {}
    missing_cols = []
    for col in _TARGET_COLUMNS:
        if col in headers:
            col_map[col] = headers.index(col)
        else:
            missing_cols.append(col)

    if missing_cols:
        return [], f"xlsx 缺少必需列: {', '.join(missing_cols)}（表头: {headers[:8]}）"

    # 逐行解析
    records = []
    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            raw_date = row[col_map["日期"]]
            raw_amount = row[col_map["金额"]]

            # 跳过空行
            if not raw_date and not raw_amount:
                continue

            # 日期处理：可能是字符串或 datetime 对象
            if hasattr(raw_date, "strftime"):
                date_str = raw_date.strftime("%Y/%m/%d %H:%M:%S")
            else:
                date_str = str(raw_date).strip() if raw_date else ""

            # 金额处理
            if isinstance(raw_amount, (int, float)):
                if raw_amount == int(raw_amount):
                    amount_str = str(int(raw_amount))
                else:
                    amount_str = str(raw_amount)
            else:
                amount_str = str(raw_amount).strip() if raw_amount else ""
                if amount_str.endswith(".0"):
                    amount_str = amount_str[:-2]

            # 类型
            raw_type = row[col_map["类型"]]
            type_str = str(raw_type).strip() if raw_type else ""

            # 分类
            raw_cat1 = row[col_map["一级分类"]]
            cat1 = str(raw_cat1).strip() if raw_cat1 else ""

            raw_cat2 = row[col_map["二级分类"]]
            cat2 = str(raw_cat2).strip() if raw_cat2 else ""

            # 备注
            raw_note = row[col_map["备注"]]
            note = str(raw_note).strip() if raw_note else ""

            record = {
                "日期": date_str,
                "类型": type_str,
                "金额": amount_str,
                "一级分类": cat1,
                "二级分类": cat2,
                "备注": note,
            }
            records.append(record)

        except Exception as e:
            _log(f"[finance.import] 行 {row_idx} 解析异常: {e}")
            continue

    _log(f"[finance.import] {filename} Sheet「{sheet_name}」解析: {len(records)} 条有效记录")
    return records, None


def _scan_and_import(ctx):
    """扫描 inbox 目录，导入所有未处理的 xlsx 文件。

    Args:
        ctx: UserContext — 提供 IO 后端和路径

    Returns:
        dict — {success, reply?, agent_context?}
    """
    # 1. 列出 inbox 目录
    items = ctx.IO.list_children(ctx.finance_inbox_dir)
    if items is None:
        return {"success": False, "reply": "无法访问 inbox 目录，请检查网络连接"}
    if not items:
        return {"success": True, "reply": "inbox 目录为空，没有新的 iCost 文件需要导入"}

    # 2. 筛选 xlsx 文件
    xlsx_files = [
        item for item in items
        if item.get("name", "").lower().endswith((".xlsx", ".xls"))
        and "file" in item  # 排除子目录
    ]
    if not xlsx_files:
        return {"success": True, "reply": "inbox 目录中没有找到 xlsx 文件"}

    # 3. 读取现有数据
    data = load_finance_data(ctx, force=True)
    if not data:
        # 初始化空数据结构
        data = {
            "version": "3.0",
            "lastModified": "",
            "data": {
                "收支账单": [],
                "资产快照": [],
                "投资持仓明细": [],
                "工资与收入": [],
                "职业里程碑": [],
                "保险规划": [],
                "每月预算": [],
            },
            "imported_files": [],
        }

    bills = data.get("data", {}).get("收支账单", [])
    imported_files = data.get("imported_files", [])

    # 4. 构建现有账单的 hash 集合（O(1) 查重）
    existing_hashes = set(_bill_hash(b) for b in bills)

    # 5. 逐个文件处理
    total_new = 0
    total_skipped = 0
    file_results = []

    for item in xlsx_files:
        fname = item["name"]

        # 检查是否已导入过
        if fname in imported_files:
            _log(f"[finance.import] 跳过已导入文件: {fname}")
            file_results.append({"file": fname, "status": "skipped", "reason": "已导入过"})
            continue

        # 下载文件
        file_path = f"{ctx.finance_inbox_dir}/{fname}"
        file_bytes = ctx.IO.download_binary(file_path)
        if file_bytes is None:
            _log(f"[finance.import] 下载失败: {fname}")
            file_results.append({"file": fname, "status": "error", "reason": "下载失败"})
            continue

        # 解析 xlsx
        records, error = _parse_xlsx_bytes(file_bytes, fname)
        if error:
            _log(f"[finance.import] 解析失败 {fname}: {error}")
            file_results.append({"file": fname, "status": "error", "reason": error})
            continue

        # 去重合并
        new_count = 0
        skip_count = 0
        for record in records:
            h = _bill_hash(record)
            if h not in existing_hashes:
                bills.append(record)
                existing_hashes.add(h)
                new_count += 1
            else:
                skip_count += 1

        total_new += new_count
        total_skipped += skip_count

        # 记录已导入
        imported_files.append(fname)

        file_results.append({
            "file": fname,
            "status": "success",
            "new": new_count,
            "skipped": skip_count,
            "total_in_file": len(records),
        })
        _log(f"[finance.import] {fname}: 新增 {new_count} 条, 跳过 {skip_count} 条重复")

    # 6. 保存数据
    if total_new > 0:
        data["data"]["收支账单"] = bills
        data["imported_files"] = imported_files
        ok = save_finance_data(ctx, data)
        if not ok:
            return {"success": False, "reply": "数据写回失败，请稍后重试"}
    else:
        # 即使没有新数据，也更新 imported_files 记录
        if any(r["status"] == "success" for r in file_results):
            data["imported_files"] = imported_files
            save_finance_data(ctx, data)

    # 7. 构建返回结果
    return {
        "success": True,
        "agent_context": {
            "files_processed": len(file_results),
            "total_new_records": total_new,
            "total_skipped_records": total_skipped,
            "total_bills_after": len(bills),
            "file_details": file_results,
        }
    }


def handle_import(params, state, ctx):
    """iCost 数据导入入口。

    params: {} — 无需参数
    """
    _log("[finance.import] 开始扫描 inbox 目录...")
    try:
        return _scan_and_import(ctx)
    except Exception as e:
        _log(f"[finance.import] 导入异常: {e}")
        import traceback
        traceback.print_exc(file=sys.stderr)
        return {"success": False, "reply": f"导入过程中出错: {str(e)}"}


# Skill 热加载注册表（V12: visibility=private，仅管理员可见可用）
SKILL_REGISTRY = {
    "finance.import": {
        "handler": handle_import,
        "visibility": "private",
        "description": "iCost 数据导入",
    },
}
